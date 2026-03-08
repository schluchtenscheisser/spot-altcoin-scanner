from pathlib import Path

from openpyxl import load_workbook

from scanner.pipeline.excel_output import ExcelReportGenerator
from scanner.pipeline.output import ReportGenerator


def _breakout_rows():
    return [
        {"symbol": "AAAUSDT", "setup_id": "breakout_immediate_1_5d", "score": 80.0, "final_score": 80.0},
        {"symbol": "BBBUSDT", "setup_id": "breakout_retest_1_5d", "score": 82.0, "final_score": 82.0},
    ]


def _trade_candidates():
    return [
        {
            "rank": 1,
            "symbol": "AAAUSDT",
            "coin_name": "AAA",
            "decision": "ENTER",
            "decision_reasons": [],
            "global_score": 80.0,
        },
        {
            "rank": 2,
            "symbol": "BBBUSDT",
            "coin_name": "BBB",
            "decision": "WAIT",
            "decision_reasons": ["entry_not_confirmed", "btc_regime_caution"],
            "global_score": 70.0,
        },
    ]


def test_pr5_markdown_contains_sot_sections() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 10}})

    md = generator.generate_markdown_report([], _breakout_rows(), [], _breakout_rows(), "2026-02-22")

    assert "## ENTER Candidates" in md
    assert "## WAIT Candidates" in md


def test_pr5_json_contains_breakout_setup_lists() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 10}})

    report = generator.generate_json_report([], _breakout_rows(), [], [], "2026-02-22")

    assert report["setups"]["breakout_immediate_1_5d"][0]["setup_id"] == "breakout_immediate_1_5d"
    assert report["setups"]["breakout_retest_1_5d"][0]["setup_id"] == "breakout_retest_1_5d"


def test_excel_has_trade_candidate_sheets(tmp_path: Path) -> None:
    generator = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})

    excel_path = generator.generate_excel_report(trade_candidates=_trade_candidates(), run_date="2026-02-22")
    wb = load_workbook(excel_path)

    assert "Trade Candidates" in wb.sheetnames
    assert "Enter Candidates" in wb.sheetnames
    assert "Wait Candidates" in wb.sheetnames


def test_pr5_markdown_renders_wait_reasons_from_trade_candidates() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 10}})

    rows = _breakout_rows()
    rows[0]["decision"] = "WAIT"
    rows[0]["decision_reasons"] = ["entry_not_confirmed", "btc_regime_caution"]
    rows[0]["global_score"] = 80.0

    md = generator.generate_markdown_report([], rows, [], rows, "2026-02-22")

    assert "decision_reasons: entry_not_confirmed, btc_regime_caution" in md


def test_excel_wait_sheet_contains_only_wait_rows(tmp_path: Path) -> None:
    generator = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})

    excel_path = generator.generate_excel_report(trade_candidates=_trade_candidates(), run_date="2026-02-22")
    wb = load_workbook(excel_path)
    ws = wb["Wait Candidates"]

    assert ws.cell(row=2, column=2).value == "BBBUSDT"
    assert ws.max_row == 2
